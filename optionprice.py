import yfinance as yf 
import numpy as np
import matplotlib.pyplot as plt 
import numpy.linalg as la 
from scipy.linalg import lu
from scipy.interpolate import interp1d
import tkinter as tk
import pandas as pd 
import pyautogui
from openpyxl import Workbook,load_workbook
import ttkbootstrap 
def autofill_fields():
	df=pd.read_excel("/Users/harshitraheja/Desktop/Tickr.xlsx")
	first_row=df.iloc[0]
	for index, rows in df.iterrows():
		entry1.delete(0,tk.END)
		entry1.insert(0,str(rows['TICKR']))
		entry2.delete(0,tk.END)
		entry2.insert(0,int(rows['Total_Exercise_Time']))
		entry3.delete(0,tk.END)
		entry3.insert(0,int(rows['expiration']))
		entry4.delete(0,tk.END)
		entry4.insert(0,int(rows['actual_exercise_duration']))
		entry5.delete(0,tk.END)
		entry5.insert(0,int(rows['Expected_price']))
		stc=entry1.get()
		ep=entry3.get()
		aed=entry4.get()
		total_exercise_time=entry2.get()
		spotprice=entry5.get()
		option_price1.set((option_price_calculator(stc,ep,aed,total_exercise_time,spotprice)))
		text1.set("Option Price")
		stock_price1.set(int(yf.Ticker(str((stc))).fast_info.last_price))
		text2.set("Current Stock Price")
		wb=load_workbook("/Users/harshitraheja/Desktop/Tickr.xlsx")
		ws=wb.active
		ws.cell(row=index+2,column=6,value=option_price1.get())
		ws.cell(row=index+2,column=7,value=stock_price1.get())
		wb.save("/Users/harshitraheja/Desktop/Tickr.xlsx")
root=tk.Tk()
stc=tk.StringVar()
ep=tk.IntVar()
aed=tk.IntVar()
total_exercise_time=tk.IntVar()
spotprice=tk.IntVar()
vol=tk.IntVar()
root.title("Call Option Price Calculator")
tk.Label(root,text="TICKR").pack()
tk.Label(root,text="Time to exercise(In days from 1st July)").pack()
tk.Label(root,text="Strike Price").pack()
tk.Label(root,text="Date of actual exercise").pack()
tk.Label(root,text="Expected price").pack()
tk.Label(root,text="Volatility").pack()
entry1=tk.Entry(root,textvariable=stc)
entry1.pack()
entry2=tk.Entry(root,textvariable=total_exercise_time)
entry3=tk.Entry(root,textvariable=ep)
entry4=tk.Entry(root,textvariable=aed)
entry2.pack()
entry3.pack()
entry4.pack()
entry5=tk.Entry(root,textvariable=spotprice)
entry5.pack()
entry6=tk.Entry(root,textvariable=vol)
entry6.pack()
def option_price_calculator(stc,ep,aed,total_exercise_time,vol,spotprice):
	volatility=vol
	r=0.045
	max_price=2*spotprice
	tet= total_exercise_time
	time_step=int(tet)
	price_step=int(tet)
	time_dif=(int(tet)/(time_step))
	price_dif=(max_price)/price_step
	v_i=0
	vi2=0
	vi1=0
	vt2=np.zeros((time_step,price_step))
	vt3=np.zeros((time_step,price_step))
	K=float(ep)
	v_i=0
	vi1=max(price_dif-K,0)
	vi2=max((2*price_dif)-K,0)
	vt2[:,-1]=max_price-K
	x=np.linspace(0,int(total_exercise_time),int(time_step))
	y=np.linspace(0,max_price,price_step)
	for j in range(0,price_step-2):
		a=(0.5*r*(price_step-j-1)*time_dif)-(0.5*(volatility**2)*((price_step-j-1)**2)*time_dif)
		b=1+((volatility**2)*((price_step-j-1)**2)*time_dif)#+(r*time_dif)
		c=(-0.5*r*(price_step-j-1)*time_dif)-(0.5*(volatility**2)*((j+1)**2)*time_dif)
		d=vt2[:-1,price_step-j-1]
		e=np.zeros((time_step-1,price_step))
		e[0,0]=a/(1-r*(time_dif))
		e[0,1]=b/(1-r*(time_dif))
		e[0,2]=c/(1-r*(time_dif))
		for l in range(1,price_step-1):
			e[l,l-1]=a/(1-r*(time_dif))
			e[l,l]=b/(1-r*(time_dif))
			e[l,l+1]=c/(1-r*(time_dif))
		g=la.pinv(e)
		f=np.transpose(e)
	
		h=np.matmul(g,d)
		vt2[:,price_step-j-2]=h
		for i in range(0,time_step-1):
			vt2[i,price_step-j-2]=max(vt2[i,price_step-j-2],max((price_step-j-2)*price_dif-K,0))
		vt2[-1,price_step-j-2]=max(((price_step-j-2)*price_dif)-K,0)
		
	def lagrange(x,y,stockprice):
		u=1
		stockprice=float(stockprice)
		for y2 in range(0,len(x)):
			u*=(stockprice-x[y2])
		def l3(x,index):
			#print(x,'x')
			w=1
			for m in range(0,len(x)):
				if m!=index:
					w*=1/(x[index]-x[m])
				else:
					w=w
			return w 
		z1=0
		for l2 in range(0,len(x)):
			if stockprice!=x[l2]:
				z1+=y[l2]*l3(x,l2)*(1/(stockprice-x[l2]))
		return z1*u
	#spotprice=yf.Ticker(str((stc))).fast_info.last_price
	Current_price=lagrange(y,vt2[int(aed),:], spotprice)
	return Current_price
stock_price1=tk.IntVar()
option_price1=tk.IntVar()
text1=tk.StringVar()
text2=tk.StringVar()
text3=tk.StringVar()
output_label = tk.Label(root, textvariable=stock_price1, font=("Arial", 14))
output_label.pack(pady=10)
output_label2 = tk.Label(root, textvariable=option_price1, font=("Arial", 14))
output_label2.pack(pady=10)
output_label3 = tk.Label(root, textvariable=text2, font=("Arial", 14))
output_label3.pack(pady=10)
output_label4 = tk.Label(root, textvariable=text1, font=("Arial", 14))
output_label4.pack(pady=10)
def submit_data():
	stc=entry1.get()
	ep=entry3.get()
	aed=entry4.get()
	spotprice=entry5.get()
	total_exercise_time=entry2.get()
	vol=float(entry6.get())
	option_price1.set((option_price_calculator(stc,ep,aed,total_exercise_time,vol,float(spotprice))))
	text1.set("Option Price")
	stock_price1.set(int(yf.Ticker(str((stc))).fast_info.last_price))
	text2.set("Current Stock Price")
button = tk.Button(root, text="Check Option Price", width=30, command=submit_data).pack()
button2 = tk.Button(root, text="Autofill", width=30, command=autofill_fields).pack()
root.mainloop()